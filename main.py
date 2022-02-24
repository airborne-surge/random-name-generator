""" Random name generator based on Social Security Administration's 'Top Names Over The Last 100 Years' and data from
of the most frequent last names from US Census Bureau from 2010 """
import contextlib
import os
import sys
import random
import urllib.request
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import *

from openpyxl import load_workbook
# from termcolor import colored
from colorama import Fore
from colorama import Style


def generate_random_name():
    """Generate Random Name"""
    names = _prepare_name_lists()

    # welcome message
    print("Welcome to the Random Name Generator.\n")

    while True:
        user_input = input("Please enter 'male' if you would like to generate a "
                           "random male name or 'female' for a female name. If you would like to exit the program, "
                           "please type 'quit'.\n>\t")
        # for termcolor package. enables color in terminal for windows systems
        # os.system('color')

        if user_input.strip().lower() == "male":
            male_name = f"{random.choice(names[0])} {random.choice(names[2])}"
            print(f" The random name is: {male_name}\n")

            # colorama version of print statement
            # print(f" The random name is: {Fore.LIGHTCYAN_EX}{male_name}{Style.RESET_ALL}\n")

            # termcolor version of print statement.
            # print(f" The random name is: {colored(male_name,'red')}\n")

        elif user_input.strip().lower() == "female":
            male_name = f"{random.choice(names[1])} {random.choice(names[2])}"
            print(f" The random name is: {male_name}\n")

        elif user_input.strip().lower() == "quit":
            print("Thank you for trying the random name generator. Now exiting...")
            sys.exit()
        else:
            print("The value you have entered is not recognized by this program. Please try again.\n")


def _prepare_name_lists():
    """Helper function checks for existence of files with first names and last names. If they exist, the list of names
    list is populated from these files. If they do not exist, this function calls on helper functions to retrieve the
    data from the SSA and US Census Bureau websites. These data is then written to a corresponding file in the
    data sets' folder. This function returns the a list that contains 3 sub-lists of the data."""

    # expected file paths if files already exist on user's system
    file_paths = ['./data sets/top_100_us_male_first_names.txt', './data sets/top_100_us_female_first_names.txt',
                  './data sets/top_100_us_last_names.txt']
    # will hold the generated names lists
    list_of_names_list = []
    # check for the existence of data files
    for path in file_paths:
        file = Path(path)
        # if a required file doesn't exist
        if not file.is_file():
            # warn the user
            print(f"A required file was not found in data folder: {os.path.basename(path)}. Fetching data from SSA "
                  f"and US Census Bureau now...")
            # get all the data necessary. put it into the main-list.
            list_of_names_list = _get_name_records()

            # write each list into corresponding file
            for index, name_list in enumerate(list_of_names_list):
                # open the file to be written to
                with open(file_paths[index], 'w') as file:
                    for line in list_of_names_list[index]:
                        # write the line
                        file.write(line + '\n')
            break
        # otherwise get list of names from file
        else:
            with open(path) as file:
                name_list = [line.rstrip() for line in file]
                list_of_names_list.append(name_list)

    # return the list
    return list_of_names_list


def _get_name_records():
    """Retrieves the first name data from the SSA site and the last name data from the US Census Bureau site. Returns
     a main list with 3 sublists: 100 male first names, 100 female first names, and 100 last names """
    # get webdriver options
    options = Options()
    # set options to headless
    options.headless = True
    # create the webdriver instance with configured options
    driver = webdriver.Chrome(options=options)

    # the site where SSA published the information
    url = "https://www.ssa.gov/oact/babynames/decades/century.html"

    # get the page
    driver.get(url)

    # wait for name table to load
    try:
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 't-stripe')))
    except TimeoutException:
        print("Unable to retrieve names. Please Try again.", file=sys.stderr)
        sys.exit()

    # find the table with the name records
    name_table = driver.find_element(By.CLASS_NAME, 't-stripe')

    # body of the table under which names are kept
    table_body = name_table.find_element(By.XPATH, '//*[@id="content"]/section[2]/div/div[2]/table/tbody')

    # the table records themselves
    name_records = table_body.find_elements(By.TAG_NAME, 'tr')

    # list to hold the names
    names = []

    # check each record for names
    for record in name_records:
        # get data from each field in the record
        name_data = record.find_elements(By.TAG_NAME, 'td')
        # for each unit of data in the list of data
        for data in name_data:
            # if the data is alphabetic (as names are)
            if data.text.isalpha():
                # insert it into the names list
                names.append(data.text)
    # print(names)
    # print(len(names))
    # separate male names from female names
    male_names = names[::2]
    female_names = names[1::2]

    # print("Male Names")
    # print(male_names)
    # print("\n")
    # print("Female Names")
    # print(female_names)
    # create a list to hold many lists
    main_names_list = [male_names, female_names, _get_last_names()]

    # return the list
    return main_names_list


def _get_last_names():
    """ Download the last names excel file from the US Census Bureau"""
    url = 'https://www2.census.gov/topics/genealogy/2010surnames/Names_2010Census_Top1000.xlsx'
    filepath = './data sets/Names_2010Census_Top1000.xlsx'
    last_names = []

    # copy the file from the US Census Bureau
    with open(filepath, 'wb') as file:
        with contextlib.closing(urllib.request.urlopen(url)) as fp:
            block_size = 1024 * 8
            while True:
                block = fp.read(block_size)
                if not block:
                    break
                file.write(block)

    # check to see if file exists
    if Path(filepath).is_file():
        # open the xlsx file
        workbook = load_workbook(filename=filepath)
        # set the current sheet as active
        sheet = workbook.active

        # row 4 to row 103 contain the first 100 last names, but only in the first column.
        for row in sheet.iter_rows(min_row=4, max_row=103, min_col=1, max_col=1, values_only=True):
            # format each name as a string rather than as a tuple
            surname = ''.join(row).title()
            # append the name to the list
            last_names.append(surname)

        # print(last_names)
        return last_names


if __name__ == "__main__":
    generate_random_name()